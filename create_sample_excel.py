#!/usr/bin/env python3
"""
创建示例Excel文件的脚本
运行此脚本将在data文件夹中生成示例Excel文件
"""

import openpyxl
from pathlib import Path

def create_excel(filename, students):
    """创建Excel文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # 设置表头
    ws['A1'] = '姓名'
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=12)
    
    # 写入学生名单
    for i, student in enumerate(students, start=2):
        ws[f'A{i}'] = student
    
    # 调整列宽
    ws.column_dimensions['A'].width = 15
    
    # 保存文件
    wb.save(filename)
    print(f'✓ 已创建: {filename}')

def main():
    # 确保data文件夹存在
    data_dir = Path('data')
    data_dir.mkdir(exist_ok=True)
    
    # 二一班学生名单
    class_2_1 = [
        '张三', '李四', '王五', '赵六', '孙七',
        '周八', '吴九', '郑十', '陈一', '刘二',
        '杨三', '黄四', '朱五', '林六', '何七',
        '高八', '徐九', '马十', '梁一', '宋二'
    ]
    
    # 二二班学生名单
    class_2_2 = [
        '王明', '李华', '张伟', '刘强', '陈静',
        '杨洋', '赵敏', '孙丽', '周杰', '吴磊',
        '郑爽', '冯巩', '卫青', '霍去病', '岳飞',
        '文天祥', '于谦', '戚继光', '郑成功', '林则徐'
    ]
    
    # 二三班学生名单
    class_2_3 = [
        '小明', '小红', '小刚', '小丽', '小华',
        '小强', '小芳', '小军', '小燕', '小东',
        '小梅', '小龙', '小凤', '小虎', '小兰',
        '小鹏', '小雪', '小松', '小云', '小峰'
    ]
    
    # 创建Excel文件
    create_excel(data_dir / '二一班.xlsx', class_2_1)
    create_excel(data_dir / '二二班.xlsx', class_2_2)
    create_excel(data_dir / '二三班.xlsx', class_2_3)
    
    print('\n✅ 所有示例Excel文件创建完成！')
    print('\n使用方法：')
    print('  - 二一班: index.html?class=二一班')
    print('  - 二二班: index.html?class=二二班')
    print('  - 二三班: index.html?class=二三班')
    print('\n提示：请使用本地服务器打开网页，而不是直接双击HTML文件')

if __name__ == '__main__':
    try:
        main()
    except ImportError:
        print('❌ 错误：未安装 openpyxl 库')
        print('\n请先安装依赖：')
        print('  pip install openpyxl')
    except Exception as e:
        print(f'❌ 错误：{e}')
